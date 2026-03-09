import json
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404, render
import openpyxl
from gestion.services import procesar_excel_stock_completo
from .models import GastoGeneral, Operacion, Producto
from django.db.models import Sum
from .forms import FacturaUploadForm
from django.shortcuts import redirect
from .utils.ia_utils import extraer_texto_pdf, procesar_con_ia
from django.contrib.auth.decorators import login_required

@login_required
def dashboard_resumen(request):
    operaciones = Operacion.objects.all()
    
    # Sumamos los montos de TODAS las operaciones cargadas
    real = operaciones.aggregate(Sum('monto_total_real'))['monto_total_real__sum'] or 0
    facturado = operaciones.aggregate(Sum('monto_facturado'))['monto_facturado__sum'] or 0
    informal = real - facturado

    return render(request, 'gestion/dashboard.html', {
        'real': real,
        'facturado': facturado,
        'informal': informal,
        'operaciones': operaciones # Para ver la lista abajo
    })

@login_required
def subir_factura(request):
    if request.method == 'POST':
        form = FacturaUploadForm(request.POST, request.FILES)
        if form.is_valid():
            operacion = form.save() 
            
            texto = extraer_texto_pdf(operacion.archivo_factura.path)
            print(f"DEBUG - Texto extraído (primeros 100 caracteres): {texto[:100]}") # [cite: 1, 9, 14]
            
            datos_ia = procesar_con_ia(texto)
            print(f"DEBUG - Respuesta de IA: {datos_ia}")
            
            if datos_ia:
                operacion.monto_facturado = datos_ia.get('monto_facturado', 0)
                operacion.es_fiscal = datos_ia.get('es_factura_valida', False)
                operacion.save()
                print(f"DEBUG - ¡Guardado con éxito! Monto: {operacion.monto_facturado}")
            else:
                print("DEBUG - La IA no devolvió nada (datos_ia es None)")
            
            return redirect('dashboard')
    else:
        form = FacturaUploadForm()
    return render(request, 'gestion/subir.html', {'form': form})

@login_required
def gestion_gastos_generales(request):
    # 1. PROCESAR LA CARGA DE NUEVOS GASTOS (POST)
    if request.method == 'POST':
        descripcion = request.POST.get('descripcion')
        categoria = request.POST.get('categoria')
        monto = request.POST.get('monto')
        tipo = request.POST.get('tipo')
        metodo_pago = request.POST.get('metodo_pago')
        fecha_manual = request.POST.get('fecha_gasto')

        if descripcion and categoria and monto and tipo and metodo_pago:
            GastoGeneral.objects.create(
                descripcion=descripcion,
                categoria=categoria,
                monto=monto,
                tipo=tipo,
                metodo_pago=metodo_pago,
                fecha=fecha_manual if fecha_manual else timezone.now()
            )
        return redirect('gastos_generales')

    # 2. CAPTURAR FILTROS (GET)
    # Por defecto traemos todos los del mes actual si no hay filtros aplicados
    gastos = GastoGeneral.objects.all().order_by('-fecha')
    
    f_desde = request.GET.get('fecha_desde')
    f_hasta = request.GET.get('fecha_hasta')
    c_filtro = request.GET.get('categoria_filtro')
    t_filtro = request.GET.get('tipo_filtro')

    # Aplicar filtros a la consulta
    if f_desde:
        gastos = gastos.filter(fecha__gte=f_desde)
    if f_hasta:
        gastos = gastos.filter(fecha__lte=f_hasta)
    if c_filtro:
        gastos = gastos.filter(categoria=c_filtro)
    if t_filtro:
        gastos = gastos.filter(tipo=t_filtro)

    # 3. LÓGICA PARA EL GRÁFICO Y RESUMEN (Basado en lo filtrado)
    # Sumamos los montos agrupados por categoría para el gráfico de dona
    resumen_query = gastos.values('categoria').annotate(total=Sum('monto'))
    
    # Diccionario para que el gráfico muestre nombres lindos
    nombres_categorias = dict(GastoGeneral.CATEGORIAS)
    
    labels = [nombres_categorias.get(item['categoria']) for item in resumen_query]
    valores = [float(item['total']) for item in resumen_query]
    
    # Calculamos el total de la selección actual para la tarjeta roja
    total_filtrado = sum(valores)

    # 4. RENDERIZADO
    return render(request, 'gestion/gastos_generales.html', {
        'gastos': gastos,
        'total': total_filtrado,
        'categorias': GastoGeneral.CATEGORIAS,
        'tipos': [('FIJO', 'Gasto Fijo'), ('VARIABLE', 'Gasto Variable')],
        'metodos_pago': GastoGeneral.METODOS_PAGO,
        # Datos para Chart.js
        'labels_js': json.dumps(labels),
        'valores_js': json.dumps(valores),
        
        # Persistencia de filtros en los inputs
        'f_desde': f_desde,
        'f_hasta': f_hasta,
        'c_filtro': c_filtro,
        't_filtro': t_filtro,
        'm_filtro': request.GET.get('metodo_pago', ''),
    })

@login_required
def exportar_gastos_excel(request):
    gastos = GastoGeneral.objects.all().order_by('-fecha')
    
    # Capturamos los parámetros
    f_desde = request.GET.get('fecha_desde')
    f_hasta = request.GET.get('fecha_hasta')
    c_filtro = request.GET.get('categoria_filtro')
    t_filtro = request.GET.get('tipo_filtro')
    m_filtro = request.GET.get('metodo_pago')

    # LIMPIEZA DE FILTROS: Solo filtramos si el valor existe y no es el string "None"
    if f_desde and f_desde != "None" and f_desde != "":
        gastos = gastos.filter(fecha__gte=f_desde)
    if f_hasta and f_hasta != "None" and f_hasta != "":
        gastos = gastos.filter(fecha__lte=f_hasta)
    if c_filtro and c_filtro != "None" and c_filtro != "":
        gastos = gastos.filter(categoria=c_filtro)
    if t_filtro and t_filtro != "None" and t_filtro != "":
        gastos = gastos.filter(tipo=t_filtro)
    if m_filtro and m_filtro != "None" and m_filtro != "":
        gastos = gastos.filter(metodo_pago=m_filtro)

    # 2. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Mensual"

    # Estilo para encabezados (Negrita y fondo gris)
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")

    headers = ['Fecha', 'Descripción', 'Categoría', 'Tipo', 'Método de Pago', 'Monto ($)']
    ws.append(headers)

    # Aplicar estilos a los encabezados
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Cargar los datos
    for gasto in gastos:
        ws.append([
            gasto.fecha.strftime('%d/%m/%Y'),
            gasto.descripcion,
            gasto.get_categoria_display(),
            gasto.get_tipo_display(),
            gasto.get_metodo_pago_display(),
            float(gasto.monto) # Importante: enviar como float para que Excel lo sume
        ])

    # Ajustar el ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    # Preparar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Gastos_{timezone.now().strftime("%m_%Y")}.xlsx"'
    wb.save(response)
    return response

@login_required
def home(request):
    return render(request, 'gestion/home.html')

@login_required
def inventario_dashboard(request):
    todos = Producto.objects.all()
    
    # Clasificación por niveles de prioridad
    criticos = [p for p in todos if p.necesita_reposicion_urgente]
    advertencias = [p for p in todos if p.necesita_reposicion() and not p.necesita_reposicion_urgente]
    stock_ok = [p for p in todos if not p.necesita_reposicion()]
    
    return render(request, 'gestion/inventario.html', {
        'criticos': criticos,
        'advertencias': advertencias,
        'stock_ok': stock_ok,
        'todos': todos
    })

@login_required
def subir_inventario_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        
        if not archivo.name.endswith('.xlsx'):
            messages.error(request, "Por favor, subí un archivo con formato .xlsx")
            return redirect('inventario_dashboard')

        try:
            creados, actualizados = procesar_excel_stock_completo(archivo)
            messages.success(request, f"¡Éxito! Se procesaron todas las páginas. Total: {creados + actualizados} productos.")
        except Exception as e:
            messages.error(request, f"Hubo un problema al procesar el Excel: {str(e)}")
            
    return redirect('inventario_dashboard')


@login_required
def actualizar_stock_manual(request, producto_id):
    if request.method == 'POST':
        producto = get_object_or_404(Producto, id=producto_id)
        nuevo_stock = request.POST.get('nuevo_stock')
        
        try:
            producto.stock_actual = int(nuevo_stock)
            producto.save()
            # No enviamos mensaje de éxito para no recargar con alertas molestas
        except (ValueError, TypeError):
            messages.error(request, "El valor ingresado no es válido.")
            
    return redirect('inventario_dashboard')